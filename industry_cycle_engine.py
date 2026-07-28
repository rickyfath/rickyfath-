"""
industry_cycle_engine.py v1.1
==============================
产业周期三重确认引擎 — 批量拉取成分股K线+季度财报, 构建行业等权价格指数, 检查三道门。

用法:
  python industry_cycle_engine.py <SW3行业名> [--gates=all|1|2]
  python industry_cycle_engine.py 半导体材料 --gates=all
  python industry_cycle_engine.py 半导体材料 --gates=1  (仅价格)

输出:
  C:/Users/Administrator/Desktop/industry_{行业名}_cycle.json

数据源:
  Gate 1 (价格): Sina K-line API → 行业等权指数
  Gate 2 (业绩): hermes_finance.py → 季度营收/净利聚合
  Gate 3 (过热): BLOCKED — 需要存货/合同负债/CAPEX/在建工程等BS+CF数据
"""

import json, sys, os, csv, urllib.request, time, subprocess
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import numpy as np
import akshare as ak

DESKTOP = r"C:\Users\Administrator\Desktop"
STOCK_DB = os.path.join(DESKTOP, "股票库", "股票库Q1.xlsx")
TODAY = date.today().isoformat()

# ============================================================
# 1. Constituent Stock Discovery
# ============================================================

def find_constituents(sw3_name, level='sw3'):
    """从股票库Q1.xlsx中找出指定行业的所有成分股。level: sw2 | sw3"""
    if not os.path.exists(STOCK_DB):
        print(f"[ERROR] xlsx not found: {STOCK_DB}")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(STOCK_DB, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    col_idx = 4 if level == 'sw3' else 3  # SW3=col4, SW2=col3
    stocks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sw_val = str(row[col_idx]).strip() if row[col_idx] else ''
        if sw_val == sw3_name:
            code = str(row[0]).strip().zfill(6)
            name = str(row[1]).strip()
            if not code.startswith('920'):
                stocks.append({'code': code, 'name': name})

    wb.close()
    level_label = 'SW2' if level == 'sw2' else 'SW3'
    print(f"[Constituents] {level_label}={sw3_name}: {len(stocks)} stocks (excl. BSE, source: 股票库Q1.xlsx)")
    return stocks, sw3_name


# ============================================================
# 2. K-line Fetch (Sina API — batch)
# ============================================================

def code_to_sina(code_6):
    if code_6.startswith(('6', '68')): return f"sh{code_6}"
    elif code_6.startswith(('0', '3', '2')): return f"sz{code_6}"
    else: return f"sh{code_6}"


def fetch_daily_kline(sina_sym, count=800):
    """拉取日线K线, 返回 [{close, date, volume}, ...]"""
    url = (
        f"https://quotes.sina.cn/cn/api/jsonp_v2.php/"
        f"data/CN_MarketDataService.getKLineData"
        f"?symbol={sina_sym}&scale=240&ma=no&datalen={count}"
    )
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://finance.sina.com.cn"
    })
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            raw = r.read().decode("utf-8")
            start = raw.find("[")
            end = raw.rfind("]")
            if start != -1 and end != -1:
                bars = json.loads(raw[start:end+1])
                # Convert to standard format
                result = []
                for b in bars:
                    try:
                        result.append({
                            'date': b.get('day', ''),
                            'close': float(b['close'])
                        })
                    except (ValueError, KeyError):
                        continue
                return result, None
            return None, "parse error"
    except Exception as e:
        return None, str(e)


def batch_fetch_kline(stocks, bars_needed=800):
    """批量拉取K线, 带进度和节流"""
    print(f"[Fetch] Pulling K-line for {len(stocks)} stocks ({bars_needed} bars each)...")
    kline_data = {}
    failed = []

    for i, s in enumerate(stocks):
        code = s['code']
        sina = code_to_sina(code)
        bars, err = fetch_daily_kline(sina, bars_needed)

        if err:
            failed.append({'code': code, 'name': s['name'], 'error': err})
            print(f"  [{i+1}/{len(stocks)}] {code} {s['name']}: FAILED ({err[:60]})")
        else:
            kline_data[code] = bars
            print(f"  [{i+1}/{len(stocks)}] {code} {s['name']}: {len(bars)} bars")

        # Throttle
        if i < len(stocks) - 1:
            time.sleep(0.15)

    print(f"  OK: {len(kline_data)} | Failed: {len(failed)}")
    return kline_data, failed


# ============================================================
# 3. Build Equal-Weight Industry Index
# ============================================================

def build_equal_weight_index(kline_data, stocks):
    """
    构建行业等权价格指数。
    每天: index = sum(all available stocks' close) / count
    归一化: 基期=100 (取最早公共日期)
    """
    # Collect all dates
    all_dates = set()
    for code, bars in kline_data.items():
        for b in bars:
            all_dates.add(b['date'])
    dates = sorted(all_dates)

    # For each date, compute equal-weight average
    index_series = []
    for d in dates:
        closes = []
        for code in kline_data:
            # Find this stock's close on this date
            stock_bars = kline_data[code]
            close = None
            for b in reversed(stock_bars):
                if b['date'] <= d:
                    close = b['close']
                    break
            if close is not None and close > 0:
                closes.append(close)

        if len(closes) >= 3:  # minimum 3 stocks for valid index point
            avg = sum(closes) / len(closes)
            index_series.append({
                'date': d,
                'value': round(avg, 4),
                'n_stocks': len(closes)
            })

    # Normalize to base=100 at earliest date
    if index_series:
        base = index_series[0]['value']
        for pt in index_series:
            pt['norm'] = round(pt['value'] / base * 100, 2)

    return index_series


# ============================================================
# 4. Rule 01 Gate 1: Price Rally Check
# ============================================================

def check_price_rally(index_series):
    """
    检查价格大涨门: 较过去756个交易日低点涨幅 >= 100%
    756个交易日 ≈ 3个日历年
    """
    if len(index_series) < 756:
        return {
            'status': 'INSUFFICIENT_DATA',
            'reason': f'Only {len(index_series)} trading days available, need >= 756'
        }

    # Take last 756 bars
    recent = index_series[-756:]
    values = [pt['value'] for pt in recent]
    low_756d = min(values)
    low_idx = values.index(low_756d)
    low_date = recent[low_idx]['date']
    current = recent[-1]['value']
    rally_pct = (current - low_756d) / low_756d * 100

    return {
        'status': 'PASS' if rally_pct >= 100 else 'FAIL',
        'current_value': round(current, 4),
        'low_756d': round(low_756d, 4),
        'low_date': low_date,
        'rally_pct': round(rally_pct, 2),
        'threshold': 100.0,
        'current_date': recent[-1]['date'],
        'n_bars': len(recent)
    }


# ============================================================
# 6. Gate 2: Earnings Surge (批量拉取+聚合季度财报)
# ============================================================

HERMES_FIN = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hermes_finance.py")
CACHE_DIR = os.path.join(DESKTOP, "股票库", "分析记录")


def fetch_one_finance(code_6):
    """Run hermes_finance.py for one stock, return parsed quarters or None"""
    cache_path = os.path.join(CACHE_DIR, f"{code_6}_finance.json")

    # Try cache first
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            qs = data.get('data', {}).get('quarters', [])
            if qs:
                return code_6, qs, 'cached'
        except:
            pass

    # Run hermes_finance.py
    try:
        r = subprocess.run(
            ['python', HERMES_FIN, code_6],
            capture_output=True, text=True, timeout=60,
            cwd=os.path.dirname(HERMES_FIN), encoding='gbk', errors='replace'
        )
        # Re-read cache
        if os.path.exists(cache_path):
            with open(cache_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            qs = data.get('data', {}).get('quarters', [])
            if qs:
                return code_6, qs, 'fetched'
    except Exception as e:
        pass

    return code_6, None, 'failed'


def batch_fetch_financials(stocks, max_workers=4):
    """并行拉取季度财报"""
    print(f"[Fetch] Pulling quarterly financials for {len(stocks)} stocks...")
    results = {}
    failed = []

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(fetch_one_finance, s['code']): s for s in stocks}
        done = 0
        for f in as_completed(futures):
            done += 1
            code, quarters, status = f.result()
            if quarters:
                results[code] = quarters
                print(f"  [{done}/{len(stocks)}] {code}: {len(quarters)} quarters ({status})")
            else:
                failed.append(code)
                print(f"  [{done}/{len(stocks)}] {code}: FAILED")

    print(f"  OK: {len(results)} | Failed: {len(failed)}")
    return results, failed


def aggregate_quarterly_earnings(finance_data):
    """
    聚合行业季度营收和净利。
    返回: [{period, total_revenue, total_np, n_stocks, rev_yoy, np_yoy}, ...]
    """
    # Collect all unique periods
    all_periods = set()
    for code, quarters in finance_data.items():
        for q in quarters:
            all_periods.add(q.get('period', ''))

    periods = sorted(all_periods)

    aggregated = []
    for period in periods:
        total_rev = 0
        total_np = 0
        n = 0
        for code, quarters in finance_data.items():
            for q in quarters:
                if q.get('period') == period:
                    rev = q.get('revenue')
                    np_val = q.get('net_profit')
                    if rev is not None and np_val is not None:
                        try:
                            total_rev += float(rev)
                            total_np += float(np_val)
                            n += 1
                        except (ValueError, TypeError):
                            pass
                    break

        if n >= 3:  # minimum 3 stocks for valid aggregate
            aggregated.append({
                'period': period,
                'total_revenue': round(total_rev, 4),
                'total_net_profit': round(total_np, 4),
                'n_stocks': n
            })

    # Compute YoY (compare same quarter one year ago)
    for i, curr in enumerate(aggregated):
        curr_period = curr['period']
        # Find same quarter last year
        try:
            parts = curr_period.split('-')
            if len(parts) >= 2:
                year = int(parts[0])
                month = parts[1]
                prev_period = f"{year-1}-{month}"
                for prev in aggregated:
                    if prev['period'].startswith(prev_period):
                        rev_yoy = (curr['total_revenue'] - prev['total_revenue']) / abs(prev['total_revenue']) * 100 if prev['total_revenue'] else None
                        np_yoy = (curr['total_net_profit'] - prev['total_net_profit']) / abs(prev['total_net_profit']) * 100 if prev['total_net_profit'] else None
                        curr['rev_yoy_pct'] = round(rev_yoy, 2) if rev_yoy is not None else None
                        curr['np_yoy_pct'] = round(np_yoy, 2) if np_yoy is not None else None
                        break
        except:
            pass

    return aggregated


def check_earnings_surge(aggregated):
    """检查业绩爆发门: 最新可得单季营收同比>=20% AND 归母净利同比>=30%"""
    # Find latest quarter with YoY data
    latest = None
    for q in reversed(aggregated):
        if q.get('rev_yoy_pct') is not None and q.get('np_yoy_pct') is not None:
            latest = q
            break

    if not latest:
        return {
            'status': 'INSUFFICIENT_DATA',
            'reason': 'No quarter with valid YoY data'
        }

    rev_ok = latest['rev_yoy_pct'] >= 20
    np_ok = latest['np_yoy_pct'] >= 30
    both_ok = rev_ok and np_ok

    return {
        'status': 'PASS' if both_ok else 'FAIL',
        'period': latest['period'],
        'revenue_yoy_pct': latest['rev_yoy_pct'],
        'net_profit_yoy_pct': latest['np_yoy_pct'],
        'threshold_rev': 20.0,
        'threshold_np': 30.0,
        'rev_ok': rev_ok,
        'np_ok': np_ok,
        'n_stocks': latest['n_stocks'],
        'total_revenue': latest['total_revenue'],
        'total_net_profit': latest['total_net_profit']
    }


# ============================================================
# ============================================================
# 7. Gate 3: Cycle Overheat (BS/CF data via akshare)
# ============================================================


def fetch_bs_cf_for_stock(code_6):
    """拉取单只股票资产负债表+现金流量表, 返回合并DataFrame"""
    sym = ('SH' if code_6.startswith(('6', '68')) else 'SZ') + code_6
    try:
        bs = ak.stock_balance_sheet_by_report_em(symbol=sym)
        cf = ak.stock_cash_flow_sheet_by_report_em(symbol=sym)

        # Standardize: use REPORT_DATE as key
        bs['REPORT_DATE'] = bs['REPORT_DATE'].astype(str).str[:10]
        cf['REPORT_DATE'] = cf['REPORT_DATE'].astype(str).str[:10]

        # Merge BS + CF on REPORT_DATE
        merged = bs.merge(cf, on='REPORT_DATE', how='inner', suffixes=('_bs', '_cf'))

        # Extract needed fields
        result = []
        for _, row in merged.iterrows():
            try:
                period = str(row['REPORT_DATE'])[:10]
                inventory = float(row['INVENTORY']) if pd.notna(row.get('INVENTORY')) else None
                contract_liab = float(row['CONTRACT_LIAB']) if pd.notna(row.get('CONTRACT_LIAB')) else None
                buy_services = float(row['BUY_SERVICES']) if pd.notna(row.get('BUY_SERVICES')) else None
                construct_asset = float(row['CONSTRUCT_LONG_ASSET']) if pd.notna(row.get('CONSTRUCT_LONG_ASSET')) else None
                cip = float(row['CIP']) if pd.notna(row.get('CIP')) else None
                total_assets = float(row['TOTAL_ASSETS']) if pd.notna(row.get('TOTAL_ASSETS')) else None

                result.append({
                    'period': period,
                    'inventory': inventory,
                    'contract_liab': contract_liab,
                    'buy_services': buy_services,
                    'capex': construct_asset,
                    'cip': cip,
                    'total_assets': total_assets
                })
            except (ValueError, TypeError, KeyError):
                continue

        return code_6, result[::-1], None  # reverse to chronological
    except Exception as e:
        return code_6, None, str(e)[:100]


def batch_fetch_bs_cf(stocks):
    """批量拉取BS/CF数据"""
    print(f"[Fetch] Pulling BS+CF data for {len(stocks)} stocks (akshare)...")
    results = {}
    failed = []

    for i, s in enumerate(stocks):
        code, data, err = fetch_bs_cf_for_stock(s['code'])
        if data:
            results[code] = data
            print(f"  [{i+1}/{len(stocks)}] {code} {s['name']}: {len(data)} quarters")
        else:
            failed.append({'code': code, 'name': s['name'], 'error': err})
            print(f"  [{i+1}/{len(stocks)}] {code} {s['name']}: FAILED ({err})")

    print(f"  OK: {len(results)} | Failed: {len(failed)}")
    return results, failed


def compute_quarterly_yoy(values_by_period, periods):
    """计算各指标同比变化率"""
    yoy = {}
    for i, period in enumerate(periods):
        parts = period.split('-')
        if len(parts) >= 2:
            prev_year = f"{int(parts[0])-1}-{parts[1]}"
            curr_vals = values_by_period.get(period)
            prev_vals = values_by_period.get(prev_year)
            if curr_vals and prev_vals:
                yoy_vals = {}
                for k in curr_vals:
                    if curr_vals[k] is not None and prev_vals.get(k) is not None and prev_vals[k] != 0:
                        yoy_vals[k] = (curr_vals[k] - prev_vals[k]) / abs(prev_vals[k])
                    else:
                        yoy_vals[k] = None
                yoy[period] = yoy_vals
    return yoy


def compute_gate3_zscore(bs_cf_data, finance_data):
    """
    Gate 3 核心计算:
    库存得分 = (存货同比Z - 合同负债同比Z + 购买商品现金支出同比Z) / 3
    产能得分 = (CAPEX同比Z + 在建工程同比Z - 资产周转率Z) / 3

    Z-score: 24季滚动窗口, 使用行业整体值(所有成分股聚合)
    """
    # Step 1: Aggregate industry-level quarterly values
    all_periods = set()
    stock_periods = {}

    for code, quarters in bs_cf_data.items():
        stock_periods[code] = {}
        for q in quarters:
            p = q['period']
            all_periods.add(p)
            stock_periods[code][p] = q

    periods = sorted(all_periods)

    # Step 2: For each period, aggregate across stocks
    industry_agg = {}
    for period in periods:
        agg = {
            'inventory': 0, 'contract_liab': 0, 'buy_services': 0,
            'capex': 0, 'cip': 0, 'total_assets': 0, 'revenue': 0, 'n': 0
        }
        for code in bs_cf_data:
            sp = stock_periods[code].get(period)
            if sp:
                for k in ['inventory', 'contract_liab', 'buy_services', 'capex', 'cip', 'total_assets']:
                    if sp.get(k) is not None:
                        agg[k] += sp[k]
                agg['n'] += 1

        if agg['n'] >= 3:
            for k in ['inventory', 'contract_liab', 'buy_services', 'capex', 'cip', 'total_assets']:
                agg[k] = agg[k]  # keep as sum (we want industry total)
            industry_agg[period] = agg

    # Step 3: Compute YoY for each indicator
    # Match same quarter last year (e.g. 2025-06-30 vs 2024-06-30)
    indicators_yoy = {}
    for period in periods:
        parts = period.split('-')
        if len(parts) >= 3:
            prev_year = f"{int(parts[0])-1}-{parts[1]}-{parts[2]}"
            curr = industry_agg.get(period)
            prev = industry_agg.get(prev_year)
            if prev is None:
                # Try fuzzy match: same month, previous year
                for p2 in periods:
                    if p2.startswith(f"{int(parts[0])-1}-{parts[1]}"):
                        prev = industry_agg.get(p2)
                        prev_year = p2
                        break
            if curr and prev and prev['n'] >= 3 and curr['n'] >= 3:
                yoy_map = {}
                for k in ['inventory', 'contract_liab', 'buy_services', 'capex', 'cip', 'total_assets']:
                    if prev[k] != 0:
                        yoy_map[k] = (curr[k] - prev[k]) / abs(prev[k])
                    else:
                        yoy_map[k] = None
                indicators_yoy[period] = yoy_map

    if len(indicators_yoy) < 8:
        return {
            'status': 'INSUFFICIENT_DATA',
            'reason': f'Only {len(indicators_yoy)} periods with YoY data, need >= 8'
        }

    # Step 4: Compute rolling Z-scores (24-quarter window)
    yoy_periods = sorted(indicators_yoy.keys())

    # Build time series for each score component
    inv_series = []   # 存货YoY
    cl_series = []    # 合同负债YoY
    bs_series = []    # 购买商品现金支出YoY
    capex_series = [] # CAPEX YoY
    cip_series = []   # 在建工程YoY
    turnover_series = []  # 资产周转率YoY (approx: rev/total_assets change)

    for period in yoy_periods:
        yoy = indicators_yoy[period]
        inv_series.append(yoy.get('inventory'))
        cl_series.append(yoy.get('contract_liab'))
        bs_series.append(yoy.get('buy_services'))
        capex_series.append(yoy.get('capex'))
        cip_series.append(yoy.get('cip'))
        # Asset turnover YoY: since revenue data is in finance_data, approximate
        # Using total_assets growth as inverse proxy (growing assets = declining turnover without rev growth)
        ta = yoy.get('total_assets')
        turnover_series.append(-ta if ta is not None else None)  # negative: asset growth without rev growth = turnover decline

    # Z-score function
    def rolling_z(series, idx, window=24):
        """Compute Z-score using rolling window of past values"""
        if series[idx] is None:
            return None
        start = max(0, idx - window + 1)
        window_vals = [v for v in series[start:idx+1] if v is not None]
        if len(window_vals) < 4:
            return None
        mean = np.mean(window_vals)
        std = np.std(window_vals)
        if std == 0:
            return 0.0
        return (series[idx] - mean) / std

    # Step 5: Compute inventory_score and capacity_score per period
    scores = {}
    for i, period in enumerate(yoy_periods):
        inv_z = rolling_z(inv_series, i)
        cl_z = rolling_z(cl_series, i)
        bs_z = rolling_z(bs_series, i)
        capex_z = rolling_z(capex_series, i)
        cip_z = rolling_z(cip_series, i)
        turn_z = rolling_z(turnover_series, i)

        inv_score = None
        cap_score = None
        if inv_z is not None and cl_z is not None and bs_z is not None:
            inv_score = (inv_z - cl_z + bs_z) / 3
        if capex_z is not None and cip_z is not None and turn_z is not None:
            cap_score = (capex_z + cip_z - turn_z) / 3

        if inv_score is not None and cap_score is not None:
            scores[period] = {
                'inventory_score': round(inv_score, 4),
                'capacity_score': round(cap_score, 4),
                'inv_z': round(inv_z, 2) if inv_z else None,
                'cl_z': round(cl_z, 2) if cl_z else None,
                'bs_z': round(bs_z, 2) if bs_z else None,
                'capex_z': round(capex_z, 2) if capex_z else None,
                'cip_z': round(cip_z, 2) if cip_z else None,
                'turn_z': round(turn_z, 2) if turn_z else None
            }

    if not scores:
        return {
            'status': 'INSUFFICIENT_DATA',
            'reason': 'No valid inventory/capacity scores computed'
        }

    # Step 6: 75th percentile threshold (use all prior periods)
    latest_period = list(scores.keys())[-1]
    all_inv = [s['inventory_score'] for s in scores.values()]
    all_cap = [s['capacity_score'] for s in scores.values()]

    inv_75 = np.percentile(all_inv, 75)
    cap_75 = np.percentile(all_cap, 75)

    # Apply: threshold = max(percentile_value, 0)
    inv_75 = max(inv_75, 0)
    cap_75 = max(cap_75, 0)

    latest = scores[latest_period]
    inv_ok = latest['inventory_score'] > inv_75
    cap_ok = latest['capacity_score'] > cap_75

    return {
        'status': 'PASS' if (inv_ok and cap_ok) else 'FAIL',
        'period': latest_period,
        'inventory_score': latest['inventory_score'],
        'inventory_75pct': round(inv_75, 4),
        'inventory_ok': inv_ok,
        'capacity_score': latest['capacity_score'],
        'capacity_75pct': round(cap_75, 4),
        'capacity_ok': cap_ok,
        'detail': latest,
        'n_periods': len(scores),
        'n_stocks_with_data': len(bs_cf_data),
        'all_scores': {k: {'inv': v['inventory_score'], 'cap': v['capacity_score']} for k, v in scores.items()}
    }


# Keep old blocked function as fallback
def check_cycle_overheat_blocked():
    """Gate 3 fallback when BS/CF unavailable"""
    return {
        'status': 'BLOCKED_BY_DATA',
        'reason': 'BS/CF data fetch failed'
    }

def check_trend_status(index_series):
    """检查当前20/200日均线状态"""
    if len(index_series) < 200:
        return {'status': 'INSUFFICIENT_DATA'}

    values = [pt['value'] for pt in index_series]
    dates = [pt['date'] for pt in index_series]

    ma20 = sum(values[-20:]) / 20
    ma200 = sum(values[-200:]) / 200
    current = values[-1]
    current_date = dates[-1]

    below_20 = current < ma20
    below_200 = current < ma200

    # Count consecutive days below 200MA
    consec_below_200 = 0
    for v in reversed(values):
        if v < ma200:
            consec_below_200 += 1
        else:
            break

    return {
        'date': current_date,
        'current': round(current, 4),
        'ma20': round(ma20, 4),
        'ma200': round(ma200, 4),
        'below_20ma': below_20,
        'below_200ma': below_200,
        'consec_below_200ma': consec_below_200,
        't200_confirmed': consec_below_200 >= 3,
        'deviation_from_20ma_pct': round((current - ma20) / ma20 * 100, 2),
        'deviation_from_200ma_pct': round((current - ma200) / ma200 * 100, 2)
    }


# ============================================================
# Main
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("Usage: python industry_cycle_engine.py <SW3行业名> [--gates=all|1|2]")
        sys.exit(1)

    sw3_name = sys.argv[1]
    gates = 'all'
    level = 'sw3'
    for arg in sys.argv[2:]:
        if arg.startswith('--gates='):
            gates = arg.split('=', 1)[1]
        elif arg.startswith('--level='):
            level = arg.split('=', 1)[1]
            if level not in ('sw2', 'sw3'):
                print(f"Invalid --level: {level}")
                sys.exit(1)

    # Step 1: Find constituents
    stocks, sw3 = find_constituents(sw3_name, level)
    if len(stocks) < 2:
        print(f"[ERROR] {sw3_name} has only {len(stocks)} non-BSE stocks, need >= 2")
        sys.exit(1)

    # Step 2: Fetch K-line
    kline_data, failed = batch_fetch_kline(stocks, bars_needed=800)

    # Step 3: Build equal-weight index
    print(f"[Build] Computing equal-weight index...")
    index_series = build_equal_weight_index(kline_data, stocks)
    print(f"  Index: {len(index_series)} daily points")
    if index_series:
        print(f"  Range: {index_series[0]['date']} -> {index_series[-1]['date']}")
        print(f"  Latest norm: {index_series[-1]['norm']:.1f} (base=100)")

    # Step 4: Gate 1
    print(f"[Gate 1] Price Rally...")
    rally = check_price_rally(index_series)
    print(f"  {rally['status']}: +{rally.get('rally_pct', '?')}% from 756d low")

    # Step 5: Rule 02 Trend
    print(f"[Trend] 20/250 MA Status...")
    trend = check_trend_status(index_series)
    if 'ma20' in trend:
        print(f"  Current: {trend['current']} | MA20: {trend['ma20']} | MA200: {trend['ma200']}")
        print(f"  Below 20MA: {trend['below_20ma']} | Below 200MA: {trend['below_200ma']}")

    # Step 6: Gate 2 (Earnings)
    earnings = {'status': 'SKIPPED', 'note': '--gates=1, skipped'}
    gate3 = {'status': 'SKIPPED'}

    if gates in ('all', '2'):
        print(f"\n[Gate 2] Earnings Surge...")
        fin_data, fin_failed = batch_fetch_financials(stocks)
        if fin_data:
            agg = aggregate_quarterly_earnings(fin_data)
            earnings = check_earnings_surge(agg)
            print(f"  {earnings['status']}: period={earnings.get('period','?')} "
                  f"rev_yoy={earnings.get('revenue_yoy_pct','?')}% "
                  f"np_yoy={earnings.get('net_profit_yoy_pct','?')}%")
            # Save aggregates for reference
            earnings['quarterly_aggregates'] = agg
            earnings['n_stocks_with_data'] = len(fin_data)
            earnings['n_stocks_failed'] = len(fin_failed)

        print(f"\n[Gate 3] Cycle Overheat...")
        bs_cf_data, bs_cf_failed = batch_fetch_bs_cf(stocks)
        if len(bs_cf_data) >= 3:
            gate3 = compute_gate3_zscore(bs_cf_data, fin_data)
            if gate3['status'] in ('PASS', 'FAIL'):
                print(f"  {gate3['status']}: inv={gate3['inventory_score']:.3f}(>{gate3['inventory_75pct']:.3f}={gate3['inventory_ok']}) "
                      f"cap={gate3['capacity_score']:.3f}(>{gate3['capacity_75pct']:.3f}={gate3['capacity_ok']}) "
                      f"period={gate3['period']}")
            else:
                print(f"  {gate3['status']}: {gate3.get('reason', '?')[:80]}")
        else:
            gate3 = check_cycle_overheat_blocked()
            print(f"  {gate3['status']}: BS/CF data available for only {len(bs_cf_data)} stocks")

    # Output
    output = {
        '_meta': {
            'industry': sw3,
            'date': TODAY,
            'engine_version': '1.1',
            'constituent_count': len(stocks),
            'kline_ok': len(kline_data),
            'kline_failed': len(failed)
        },
        'rule_01': {
            'gate_1_price_rally': rally,
            'gate_2_earnings_surge': earnings,
            'gate_3_cycle_overheat': gate3,
            'triple_gate_triggered': (
                rally.get('status') == 'PASS' and
                earnings.get('status') == 'PASS' and
                gate3.get('status') == 'PASS'
            )
        },
        'rule_02_trend': trend,
        'constituents': [{'code': s['code'], 'name': s['name']} for s in stocks],
        'index_summary': {
            'n_points': len(index_series),
            'start_date': index_series[0]['date'] if index_series else None,
            'end_date': index_series[-1]['date'] if index_series else None,
            'latest_norm': index_series[-1]['norm'] if index_series else None
        }
    }

    safe_name = sw3.replace('/', '_').replace('\\', '_')
    out_path = os.path.join(DESKTOP, f"industry_{safe_name}_cycle.json")
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n[SAVED] -> {out_path}")


if __name__ == '__main__':
    main()
